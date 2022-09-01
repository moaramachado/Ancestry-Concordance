'''This module contains functions for getting the validation and original files from S3, do some format changes, run concordance and generate a table with results.'''
import boto3
import os
import re
import numpy as np
import scipy.stats
import subprocess
import sys
import pandas as pd
import xlsxwriter
import openpyxl
import time

from pathlib import Path
from typing import Any, Dict, Union, Optional

PathLike = Union[str, Path]

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Border, Side


#Creating a global variable to store the folder information for the validation batch.
#This folder will be found inside the function get_demux_stats and it will be used inside of the function find_validation_runs.
global_folder = ""


def get_demux_stats(
        prod_ac: str,
        batch: str,
        output_path: PathLike,
        bucket_name: str = 's4-ancestry-processor-incoming0b397865-1nzskcxw9agup',
        dir_path: PathLike = 'mapping/'
) -> None:
    """Downloads a single demux stats file associated with a batch to a given output location.
    Args:
        prod_ac (str): AWS profile for the germline-prod account
        batch (str): Batch ID for the validation run
        output_path (PathLike): Output path where demux stats will be downloaded to.
        bucket_name (str): The S3 bucket where the samples will be searched. Defaults to 's4-ancestry-processor-incoming0b397865-1nzskcxw9agup'.
        dir_path (PathLike, optional): S3 prefix path, Defaults to 'mapping/'.
    """

    #Reading the credentials from your germline-production account profile
    session = boto3.Session(profile_name=prod_ac)
    client = session.client('s3')
    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)

    #Finding the folder based on the batch ID information
    cmd = ('aws s3 ls s3://' + bucket_name + '/' + dir_path + ' | grep ' + batch + ' | awk \'{print $2}\'')
    folder = subprocess.getoutput(cmd)
    f = str(folder)
    print(f)

    #Changing the global_folder variable
    global global_folder
    global_folder = subprocess.getoutput(cmd)

    sub_path = str(Path(dir_path) / f) + "/"
    print(sub_path)
    #print(output_path)

    #Checking if the Ancestry_results folder exist to store demux file
    mydir = (f'{output_path}/Ancestry_results/')
    CHECK_FOLDER = os.path.isdir(mydir)

    # If folder doesn't exist, then create it.
    if not CHECK_FOLDER:
        os.makedirs(mydir)
        print("created folder : ", mydir)

    else:
        print(mydir, "folder already exists.")

    try:

        result = client.list_objects(Bucket=bucket_name, Prefix=sub_path, Delimiter='/')

        #Finding the demux file and downloading it
        s3 = session.resource('s3')
        bucket = s3.Bucket(bucket_name)

        for object_summary in bucket.objects.filter(Prefix=sub_path):
            if object_summary.key.endswith('Undetermined.csv'):
                demux_file = object_summary.key
                #print(demux_file)
                file_name = demux_file.split("/", 3)
                demux_name = file_name[2]
                print("Demux file:", demux_name)
                bucket.download_file(demux_file, str(f'{output_path}/Ancestry_results/{demux_name}'))

    except ValueError:
        raise


def find_validation_runs(
    test_ac: str,
    v_batch: str,
    output_path: PathLike,
    samples: str,
    bucket_name: str = 's4-ancestry-processor-incoming0b397865-1rym98novarq5',
    dir_path: PathLike = 'mapping/',
    original_batch: Optional[str] = None
) -> None:
    """Finds the validation run based on the batch ID and downloads the ancestry results for the samples (all batch or selected samples).
    Args:
        test_ac (str): AWS profile for the germline-test account.
        v_batch (str): Batch ID for the validation run.
        output_dir (PathLike): output directory path
        samples (str, None): Samples from validation batch to include.
        bucket_name (str): The S3 bucket where the samples will be searched. Defaults to 's4-ancestry-processor-incoming0b397865-1rym98novarq5'.
        original_batch (str, optional): Original batch to restrict results to. Defaults to None.
    """

    #Reading the credentials for test account
    print("Account:", test_ac)
    session = boto3.Session(profile_name=test_ac)
    client = session.client('s3')
    #client = boto3.client('s3')
    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)

    #Global folder
    folder = global_folder
    f = str(folder)
    print(f)

    sub_path = str(Path(dir_path) / f) + "/"
    print(sub_path)
    print("Output path: ", output_path)
    #sub_path = str(Path(dir_path) / v_batch) + '/'

    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)
    id_list = []
    print("Bucket:", bucket)

    #Subset if sample list has been passed in
    if samples is not None:
        sample_selection = re.sub(r'\s+', '', samples).replace('V-', '').replace(',', '|')
        print(f'Extracting sample subset: {sample_selection}')
        subset_list = sample_selection.split("|")

        #Using the samples passed to find the files and create the id_list
        for id in subset_list:
            for object_summary in bucket.objects.filter(Prefix=sub_path):
                if (object_summary.key.endswith('R1_001.fastq.gz')):
                    val_samples = object_summary.key
                    #print(val_samples)
                    samples = val_samples.split("/", 6)
                    if (samples[5].startswith(id)):
                        id_list.append(samples[5])
                        #print(id_list)

    #Get the sample list for the whole batch
    else:
        for object_summary in bucket.objects.filter(Prefix=sub_path):
            # print(object_summary)
            if object_summary.key.endswith('R1_001.fastq.gz'):
                val_samples = object_summary.key
                # print(val_samples)
                samples = val_samples.split("/", 6)
                id_list.append(samples[5])
                #print(id_list)

    #Checking if the Validation_files folder exist to store the validation samples
    mydir = (f'{output_path}/Ancestry_results/Validation_files')
    CHECK_FOLDER = os.path.isdir(mydir)

    # If folder doesn't exist, then create it.
    if not CHECK_FOLDER:
        os.makedirs(mydir)
        print("created folder : ", mydir)

    else:
        print(mydir, "folder already exists.")

    #Downloading the ancestry results based on the list of samples created (it will change the bucket - germline-test account)
    bucket_name = 's4-ancestry-processor-outgoingc68740cc-1i9tav35evnt0'
    sub_path = 'ancestry_high/'
    client = session.client('s3')
    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)

    #Creating a file with the samples IDs downloaded
    f = open(f'{output_path}/Ancestry_results/Validation_files/val_samples_id.txt', 'w')

    #Find samples from id_list in the bucket
    for files in bucket.objects.filter(Prefix=sub_path):
        #print(files.key)
        for i in id_list:
            #print (i)
            if files.key.startswith(f'ancestry_high/{i}'):
                #print(files.key)
                val_samples = files.key.split("/",2)
                #Creating a file name and downloading the sample
                name_file = val_samples[1]
                bucket.download_file(files.key, str(f'{output_path}/Ancestry_results/Validation_files/{name_file}'))

                #Creating a list of samples downloaded to be the reference for the original samples search
                val_samples_id = val_samples[1].split("_",2)
                print(val_samples_id[0], file=f)


def find_original_runs(
    prod_ac: str,
    v_batch: str,
    output_path: PathLike,
    samples: str,
    bucket_name: str = 's4-ancestry-processor-incoming0b397865-1nzskcxw9agup',
    dir_path: PathLike = 'mapping/',
    original_batch: Optional[str] = None
) -> pd.DataFrame:
    """Finds the validation run based on the batch ID and downloads the ancestry results for the samples (all batch or selected samples).

    Args:
        prod_ac (str): AWS profile for the germline-prod account.
        v_batch (str): Batch ID for original samples.
        output_dir (PathLike): output directory path.
        samples (str, None): Samples to include.
        bucket_name (str): The S3 bucket to search.
        original_batch (str, optional): Original batch to restrict results to. Defaults to None.
    """

    session = boto3.Session(profile_name=prod_ac)
    client = session.client('s3')
    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)

    #Finding the folder based on the batch ID information
    cmd = ('aws s3 ls s3://' + bucket_name + '/' + dir_path + ' | grep ' + v_batch + ' | awk \'{print $2}\'')

    folder = subprocess.getoutput(cmd)
    f = str(folder)
    print(f)

    sub_path = str(Path(dir_path) / f) + "/"
    print(sub_path)
    print(output_path)

    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)
    id_list = []

    #Subset if sample list has been passed in. It will create the list of samples to be downloaded in id_list
    if samples is not None:
        sample_selection = re.sub(r'\s+', '', samples).replace('V-', '').replace(',', '|')
        print(f'Extracting sample subset: {sample_selection}')
        subset_list = sample_selection.split("|")

        for id in subset_list:
            medgis_id = id.split("_") #Selecting only the ID from the samples (removing extracting and repetition information)
            print(medgis_id[0])
            for object_summary in bucket.objects.filter(Prefix=sub_path):
                if (object_summary.key.endswith('R1_001.fastq.gz')):
                    ori_samples = object_summary.key
                    samples = ori_samples.split("/", 6)
                    if (samples[5].startswith(medgis_id[0])):
                        id_list.append(samples[5])
                        #print(id_list)

    #Get all samples from the batch
    else:
        for object_summary in bucket.objects.filter(Prefix=sub_path):
            #print(object_summary)
            if object_summary.key.endswith('R1_001.fastq.gz'):
                ori_samples = object_summary.key
                #print(ori_samples)
                samples = ori_samples.split("/", 6)
                id_list.append(samples[5])
                #print(id_list)

    #Checking if the Original_files folder exist to store the original samples
    mydir = (f'{output_path}/Ancestry_results/Original_files')
    CHECK_FOLDER = os.path.isdir(mydir)

    # If folder doesn't exist, then create it.
    if not CHECK_FOLDER:
        os.makedirs(mydir)
        print("created folder : ", mydir)

    else:
        print(mydir, "folder already exists.")

    #Downloading the ancestry results based on the list of samples created (it will change the bucket - germline-prod account)
    bucket_name = 's4-ancestry-processor-outgoingc68740cc-1pz98jf0gqubn'
    sub_path = 'ancestry_high/'
    client = session.client('s3')
    s3 = session.resource('s3')
    bucket = s3.Bucket(bucket_name)

    f = open(f'{output_path}/Ancestry_results/Original_files/original_samples_id.txt', 'w')

    for files in bucket.objects.filter(Prefix=sub_path):
        for i in id_list:
            if files.key.startswith(f'ancestry_high/{i}'):
                print(files.key)
                ori_samples = files.key.split("/",2)
                #Creating a file name and downloading the samples
                name_file = ori_samples[1]
                bucket.download_file(files.key, str(f'{output_path}/Ancestry_results/Original_files/{name_file}'))

                #Creating a list of samples downloaded to be the reference for the original samples search
                ori_samples_id = ori_samples[1].split("_",2)
                print(ori_samples_id[0], file=f)


def generate_validation_csv_files(
    output_path: PathLike,
) -> pd.DataFrame:
    """Generates csv files from the json files that were downloaded.
    Args:
        output_dir (PathLike): output directory path.
    """

    sub_path = str(f'{output_path}/input/Validation_files')
    #print(sub_path)
    print("Output pathway:", output_path)

    input_directory = sub_path
    out_directory = str(output_path) + '/' + 'input/Validation_files' + '/' + 'json2csv_validation'

    ##Part1: Create the output folder with the csv files for each sample
    if not os.path.exists(out_directory):
        os.mkdir(out_directory)
        print("created folder : ", out_directory)

    else:
        print(out_directory, "folder already exists.")


    for filename in os.listdir(input_directory):
        if filename.endswith(".json"):
            f = open(os.path.join(input_directory, filename))
            df = pd.read_json(f)

            ##Taking care of NaN values inside DataFrame and replacing them with 0
            df.fillna(0)

            key_list = list(df.iloc[0, 0].keys())
            col_list = list(df.columns)
            row_list = list(df.index)

            ss_arr = np.zeros(shape=(df.shape[0], len(col_list)))
            df_final = pd.DataFrame(ss_arr, index=row_list, columns=col_list)

            for i in range(0, len(row_list)):
                for j in range(0, len(col_list)):
                    if type(df.iloc[i, j]) is dict and "mean" in df.iloc[i, j]:
                        df_final.iloc[i, j] = df.iloc[i, j]['mean']

            df_final.to_csv(out_directory + '/' + os.path.splitext(filename)[0] + '.csv', encoding='utf-8', index=True)

    print()

    ##Part2: Combine all the csv files
    sub_path2 = str(f'{output_path}/input/Validation_files/json2csv_validation')
    print("Validation files:", sub_path2)
    print("Output path:", output_path)

    input_directory = sub_path2
    out_directory = str(output_path) + '/' + 'input/Validation_files' + '/' + 'json2csv_validation'



    # Input folder for the csv files that was generated in Part1
    source_files = sorted(Path(f'{out_directory}').glob('*.csv'))

    dataframes = []
    for file in source_files:
        df = pd.read_csv(file)  # additional arguments up to your needs
        df['source'] = file.name
        dataframes.append(df)

    df_all = pd.concat(dataframes, sort=True)

    ##Removing duplicated information
    df_with_index = df_all.set_index("Unnamed: 0")

    #dropping passed values
    df_with_index = df_with_index.drop("gencove")

    #Extracting sample name
    df_with_index['source'] = df_with_index['source'].str.split('_A').str[0]

    #Creating the columns for ancestry groups that are missing
    cols = ['AFR', 'AMBIGUOUS', 'AMR', 'ASJ', 'EAS', 'FIN', 'NFE', 'SAS', 'UNASSIGNED']
    df_with_index = df_with_index.reindex(sorted(df.columns.union(cols, sort=False)), axis=1, fill_value='0.0')
    df_with_index = df_with_index.drop('Unnamed: 0', axis=1, inplace=False)

    #Printing results in a file
    df_with_index.to_csv(out_directory + '/' + 'combined_csv.tsv', index=False, sep="\t")
    print()


def generate_original_csv_files(
    output_path: PathLike,
) -> pd.DataFrame:
    """Generates csv files from the json files that were downloaded.
    Args:
        output_dir (PathLike): output directory path.
    """

    sub_path = str(f'{output_path}/input/Original_files')
    print("Original files:", sub_path)
    print("Output path:", output_path)

    input_directory = sub_path
    out_directory = str(output_path) + '/' + 'input/Original_files' + '/' + 'json2csv_original'

    ##Part1: Create the output folder with the csv files for each sample
    if not os.path.exists(out_directory):
        os.mkdir(out_directory)
        print("created folder : ", out_directory)
    else:
        print(out_directory, "folder already exists.")


    for filename in os.listdir(input_directory):
        if filename.endswith(".json"):
            f = open(os.path.join(input_directory, filename))
            df = pd.read_json(f)

            ##Taking care of NaN values inside DataFrame and replacing them with 0
            df.fillna(0)

            key_list = list(df.iloc[0, 0].keys())
            col_list = list(df.columns)
            row_list = list(df.index)

            ss_arr = np.zeros(shape=(df.shape[0], len(col_list)))
            df_final = pd.DataFrame(ss_arr, index=row_list, columns=col_list)

            for i in range(0, len(row_list)):
                for j in range(0, len(col_list)):
                    if type(df.iloc[i, j]) is dict and "mean" in df.iloc[i, j]:
                        df_final.iloc[i, j] = df.iloc[i, j]['mean']

            df_final.to_csv(out_directory + '/' + os.path.splitext(filename)[0] + '.csv', encoding='utf-8', index=True)

    print()

    ##Part2: Combine all the csv files
    sub_path2 = str(f'{output_path}/input/Original_files/json2csv_original')

    input_directory = sub_path2
    out_directory = str(output_path) + '/' + 'input/Original_files' + '/' + 'json2csv_original'


    # Input folder for the csv files that was generated in the Part1
    source_files = sorted(Path(f'{out_directory}').glob('*.csv'))

    dataframes = []
    for file in source_files:
        df = pd.read_csv(file)  # additional arguments up to your needs
        df['source'] = file.name
        dataframes.append(df)

    df_all = pd.concat(dataframes, sort=True)

    ##Removing duplicated information
    #data with index
    df_with_index = df_all.set_index("Unnamed: 0")

    # dropping passed values
    df_with_index = df_with_index.drop("gencove")

    #Extracting sample name
    df_with_index['source'] = df_with_index['source'].str.split('_A').str[0]

    # Creating the columns for ancestry groups that are missing
    cols = ['AFR', 'AMBIGUOUS', 'AMR', 'ASJ', 'EAS', 'FIN', 'NFE', 'SAS', 'UNASSIGNED']
    df_with_index = df_with_index.reindex(sorted(df.columns.union(cols, sort=False)), axis=1, fill_value='0.0')

    df_with_index = df_with_index.drop('Unnamed: 0', axis=1, inplace=False)


    #Printing results in a file
    df_with_index.to_csv(out_directory + '/' + 'combined_csv.tsv', index=False, sep="\t")
    print()


def run_concordance(
    output_path: PathLike,
    original_batch: Optional[str] = None
) -> pd.DataFrame:
    """Merges validation and original files and run concordance.
    Args:
        output_dir (PathLike): output directory path.
    """

    sub_path_validation = str(f'{output_path}/input/Validation_files/json2csv_validation')
    sub_path_original = str(f'{output_path}/input/Original_files/json2csv_original')

    out_file = str(f'{output_path}/input/ancestry.tsv')

    file_validation = str(f'{sub_path_validation}/combined_csv.tsv')
    file_original = str(f'{sub_path_original}/combined_csv.tsv')

    df1 = pd.read_csv(file_validation, delimiter='\t')
    df2 = pd.read_csv(file_original, delimiter='\t')

    df1['id'] = df1['source'].str[:8]
    df2['id'] = df2['source'].str[:8]
    #df = pd.DataFrame(df1['source'].tolist())

    file_merged = df1.merge(df2, indicator=True, how='outer', on='id')

    #Printing results in a file
    file_merged.to_csv(out_file, index=False, sep="\t")

    ##Part2: calculating the concordance
    data = pd.read_csv(out_file, sep='\t')

    #Check if validation and reference groups are in the correct position
    a = data.iloc[:, 0:9]
    b = data.iloc[:, 11:20]

    #Filling the NaN with zeros
    data = data.fillna(0)

    #Printing results in a file
    data.to_csv(str(output_path) + '/' + 'input/all_ancestries.tsv', index=False, sep="\t")

    print('File all_ancestries.tsv is done!')
    print()

    results = open(str(output_path) + '/' + 'input' + '/' + 'concordance.tsv', 'w')
    results.write("id\tCorrelation(r,p-value)\n")

    #Calculating correlation
    n = len(a)
    for i in range(0, n):
        #print (i)
        id = str(data.iloc[i, 9])
        id2 = id[0:8]
        cor = scipy.stats.pearsonr(data.iloc[i, 0:9], data.iloc[i, 11:20])
        #print(cor)
        print(id[0:8], cor)
        results.write("%s\t%s\n" % (id2, cor))

    results.close()

    #Merging the ancestry file with the correlation results file based on the id column
    ancestry_file = str(f'{output_path}/input/all_ancestries.tsv')
    ancestry = pd.read_csv(ancestry_file, delimiter='\t')

    cor_file = str(f'{output_path}/input/concordance.tsv')
    correlation = pd.read_csv(cor_file, delimiter='\t')

    file2_merged = ancestry.merge(correlation, how='outer', on='id')
    file2_merged = file2_merged.dropna(subset=["_merge"]) #removing empty rows

    final_results = file2_merged[['source_x', 'AFR_x', 'AMBIGUOUS_x', 'AMR_x', 'ASJ_x', 'EAS_x', 'FIN_x', 'NFE_x', 'SAS_x','UNASSIGNED_x', 'source_y', 'AFR_y', 'AMBIGUOUS_y', 'AMR_y', 'ASJ_y', 'EAS_y', 'FIN_y', 'NFE_y', 'SAS_y', 'UNASSIGNED_y', 'Correlation(r,p-value)']]

    final_results.set_axis(['Samples', 'AFR', 'AMBIGUOUS', 'AMR', 'ASJ', 'EAS', 'FIN', 'NFE', 'SAS', 'UNASSIGNED', 'Samples', 'AFR',
         'AMBIGUOUS', 'AMR', 'ASJ', 'EAS', 'FIN', 'NFE', 'SAS', 'UNASSIGNED', 'Correlation (r, p-value)'], axis=1, inplace=True)

    #Inserting extra column to have space between validation and reference samples
    final_results.insert(10, "", " ")
    final_results.to_excel(str(output_path) + '/' + 'ancestry_concordance.xlsx', index=False)

    #Making changes in the format for the excel file with the results
    file_name = (str(output_path) + '/' + 'ancestry_concordance.xlsx')

    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    #Adding a row on top to create another header
    sheet.insert_rows(1)
    sheet.merge_cells('A1:J1')
    # sheet.cell(row = 1, column = 1).value = 'Validation samples'

    #Adding header and changing format for validation samples (merging cells and coloring)
    top_left_cell = sheet['A1']
    top_left_cell.value = "Validation samples"
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin", color="000000")

    top_left_cell = sheet['A1']
    top_left_cell.value = "Validation samples"
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
    top_left_cell.fill = PatternFill("solid", fgColor="FFF3AA")
    top_left_cell.font = Font(b=True, size=13, color="000000")
    top_left_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #Adding header and changing format for Reference samples (merging the cells and coloring)
    sheet.merge_cells('L1:U1')
    # sheet.cell(row = 1, column = 11).value = 'Reference samples'
    top_left_cell = sheet['L1']
    top_left_cell.value = "Reference samples"
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
    top_left_cell.fill = PatternFill("solid", fgColor="DDE8A6")
    top_left_cell.font = Font(b=True, size=13, color="000000")

    #Function to add border to the cells in the excel file according to the size of the file
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    rows = sheet.max_row
    range_1 = "A1:V"
    range_2 = str(rows)
    range_defined = range_1 + range_2
    print(range_defined)
    set_border(sheet, range_defined)

    #Final results saved in the ancestry_concordance.xlsx file
    workbook.save(str(output_path) + '/' + "ancestry_concordance.xlsx")
    os.remove((str(output_path) + '/' + "input" + '/' + "ancestry.tsv"))

    print("Done!")
    print("ancestry_concordance.xlsx file was created")


#def run_ancestry(prod_ac, test_ac, validation, original, pathway, samples):
#    print(get_demux_stats(prod_ac, validation, pathway))
#    print(find_validation_runs(test_ac, validation, pathway, samples))
#    print(find_original_runs(prod_ac, original, pathway, samples))
#    print(generate_validation_csv_files(pathway))
#    print(generate_original_csv_files(pathway))
#    print(run_concordance(pathway))


def main():
    generate_validation_csv_files(sys.argv[1])
    generate_original_csv_files(sys.argv[1])
    run_concordance(sys.argv[1])

if __name__ == "__main__":
    main()
